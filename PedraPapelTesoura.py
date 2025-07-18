import random  # Biblioteca para gerar escolhas aleatórias do computador
import os      # Biblioteca para limpar a tela no terminal (cross-platform)
from datetime import datetime  # Biblioteca para capturar data e hora atuais
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Dicionário com as opções do jogo
opcoes = {
    1: "✊ Pedra",
    2: "✋ Papel",
    3: "✌️  Tesoura"
}

# Variáveis para armazenar a pontuação
placar_jogador = 0
placar_computador = 0

# Função para limpar a tela (Windows ou Linux/Mac)
def limpar_tela():
    os.system('cls' if os.name == 'nt' else 'clear')

# Função para registrar a partida na planilha
def registrar_partida_xlsx(nome_jogador, placar_jogador, placar_computador):
    # Padroniza o nome do jogador (ex: leo → Leo)
    nome = nome_jogador.strip().capitalize()

    # Define o nome do arquivo
    arquivo = "registro_completo.xlsx"

    # Se o arquivo já existe, abrimos ele. Caso contrário, criamos um novo Workbook
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
    else:
        wb = Workbook()
        # Remove a aba padrão chamada "Sheet"
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ------------------------
    # Aba 1: Registro de Partidas
    # ------------------------

    # Se a aba não existir ainda, criamos ela e adicionamos cabeçalhos
    if "RegistroPartidas" not in wb.sheetnames:
        aba_partidas = wb.create_sheet("RegistroPartidas")
        aba_partidas.append(["Data", "Jogador", "Placar Jogador", "Placar Computador"])
    else:
        aba_partidas = wb["RegistroPartidas"]

    # Adiciona a linha de registro com data atual
    data = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    aba_partidas.append([data, nome, placar_jogador, placar_computador])

    # ------------------------
    # Aba 2: Ranking de Jogadores
    # ------------------------

    # Se a aba não existir, criamos ela com cabeçalho
    if "RankingJogadores" not in wb.sheetnames:
        aba_ranking = wb.create_sheet("RankingJogadores")
        aba_ranking.append(["Jogador", "Vitórias", "Derrotas", "Empates"])
    else:
        aba_ranking = wb["RankingJogadores"]

    # Verifica se o jogador já existe na tabela
    jogador_encontrado = False
    for linha in aba_ranking.iter_rows(min_row=2, values_only=False):  # min_row=2 pula o cabeçalho
        celula_nome = linha[0]
        if celula_nome.value and celula_nome.value.strip().capitalize() == nome:
            jogador_encontrado = True
            # Atualiza vitórias ou derrotas
            if placar_jogador > placar_computador:
                linha[1].value += 1  # Vitórias
            elif placar_jogador < placar_computador:
                linha[2].value += 1  # Derrotas
            elif placar_jogador == placar_computador:
                linha[3].value += 1 # Empates
            break

    # Se o jogador ainda não existir, cria nova linha
    if not jogador_encontrado:
        vitoria = 1 if placar_jogador > placar_computador else 0
        derrota = 1 if placar_computador > placar_jogador else 0
        empate = 1 if placar_jogador == placar_computador else 0
        aba_ranking.append([nome, vitoria, derrota, empate])

    # Salva tudo no arquivo
    wb.save(arquivo)

# Solicita o nome do jogador
nome_jogador = input("Digite seu nome para começar: ")

# Função que exibe o placar (parcial ou final, dependendo do argumento)
def exibir_placar(tipo):
    print(f"\nPlacar {tipo}:")
    print(f"{nome_jogador}: {placar_jogador} | Computador: {placar_computador}")

# Loop principal do jogo
while True:
    limpar_tela()  # Limpa a tela a cada rodada

    # Exibe o menu de jogadas
    print("\nEscolha sua jogada:")
    print("1 - ✊ Pedra")
    print("2 - ✋ Papel")
    print("3 - ✌️  Tesoura")
    print("9 - ❌ Sair do jogo")

    # Mostra o placar parcial antes da rodada
    exibir_placar("parcial")

    # Valida a entrada do jogador
    while True:
        jogador = input("\nDigite o número da sua jogada: ")
        if jogador in ["1", "2", "3", "9"]:
            jogador = int(jogador)
            break
        else:
            print("❌ Jogada inválida. Tente novamente.")

    # Se o jogador quiser sair (9), exibe o placar final
    if jogador == 9:
        print(f"\nObrigado por jogar {nome_jogador}! Até a próxima 👋")
        exibir_placar("final")

        # Exibe quem venceu a sessão
        if placar_jogador == placar_computador:
            print("\nO jogo terminou empatado!")
        elif placar_jogador > placar_computador:
            print("\nParabéns! Você ganhou!")
        else:
            print("\nVocê perdeu! Melhor sorte na próxima!")

        # Pergunta se deseja salvar o resultado em arquivo
        salvar = input("\nDeseja salvar esta partida e atualizar o ranking? (s/n): ").lower()
        if salvar == "s":
            registrar_partida_xlsx(nome_jogador, placar_jogador, placar_computador)
            print("✅ Partida e ranking atualizados com sucesso!")
        else:
            print("📁 Registro não salvo.")
        break # Sai do loop principal

    # Computador faz uma jogada aleatória
    computador = random.choice([1, 2, 3])

    # Exibe as escolhas
    print(f"\n{nome_jogador} escolheu: {opcoes[jogador]}")
    print(f"O computador escolheu: {opcoes[computador]}")

    # Lógica de vitória, empate ou derrota
    if jogador == computador:
        print("\n🤝 Deu empate!")
    elif (jogador == 1 and computador == 3) or \
         (jogador == 2 and computador == 1) or \
         (jogador == 3 and computador == 2):
        print(f"\n🏆 {nome_jogador} vence!")
        placar_jogador += 1
    else:
        print("\n💻 Computador vence!")
        placar_computador += 1

    # Aguarda o jogador pressionar Enter para continuar
    input("\nPressione Enter para continuar...")
